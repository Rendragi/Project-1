import pandas as pd 
from openpyxl import load_workbook 
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import datetime
import datetime as dt
import string
import logging


class ExcelReportPlugin():
    def __init__(self,
                 input_file,
                 output_file
                 ):
        self.input_file = input_file
        self.output_file = output_file

    def main(self):
        df = self.read_input_file()
        df_transform = self.transform(df)
        self.create_output_file(df_transform)
        print("workbook created")

        wb = load_workbook(self.output_file)
        ws = wb['Report']

        min_column = ws.min_column
        max_column = ws.max_column
        min_row = ws.min_row
        max_row = ws.max_row

        # Mengubah format tanggal di sel Date
        date_format = 'DD-MM-YYYY'  
        for cell in ws['A']:
            cell.number_format = date_format
            if isinstance(cell.value, dt.datetime):
                cell.value = cell.value.date()

        # Menambahkan border ke seluruh tabel dalam range
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=min_row, min_col=min_column, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        self.column_dimension(ws)
        self.barchart1(ws, min_column, max_column, min_row, max_row)
        self.barchart2(ws, min_column, max_column, min_row, max_row)
        self.total_revenue_perproduct(max_column, max_row, min_row, ws)
        self.add_title(ws)
        self.save_file(wb)

    def read_input_file(self):
        df = pd.read_excel(self.input_file)
        logging.info(df.head())
        return df

    #Transform Data
    def transform(self, df:pd.DataFrame) -> pd.DataFrame:
        df_transform = df.pivot_table(index='Date', 
                                    columns=['Gender', 'Product line'],  
                                    values='gross income', 
                                    aggfunc='sum').round(2)
        return df_transform
    
    #Membuat output File
    def create_output_file(self, df):
        print('Save dataframe to excel...')
        df.to_excel(self.output_file, 
                        sheet_name='Report', 
                        startrow=4)
        print(f'Save dataframe done... {self.output_file}')

    #Atur Dimensi kolom
    def column_dimension(self, workbook):
        dim_holder = DimensionHolder(worksheet=workbook)

        for col in range(workbook.min_column, workbook.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(workbook, min=col, max=col, width=20)

        workbook.column_dimensions = dim_holder

    #Membuat Barchart untuk Female
    def barchart1(self, workbook, min_column, max_column, min_row, max_row):
        barchart = BarChart()

        data = Reference(workbook, 
                        min_col=min_column+1,
                        max_col=min_column+6,
                        min_row=min_row+1,
                        max_row=max_row
                        )

        categories = Reference(workbook,
                                min_col=min_column,
                                max_col=min_column,
                                min_row=min_row+3,
                                max_row=max_row
                                )

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categories)


        workbook.add_chart(barchart, 'C12')
        barchart.title = 'Daily Gross Revenue Female'
        barchart.type = 'col'
        barchart.grouping = 'stacked'
        barchart.overlap = 100
        barchart.style = 2

    #Membuat Barchart untuk Male
    def barchart2(self, workbook, min_column, max_column, min_row, max_row):
        barchart = BarChart()

        data = Reference(workbook, 
                        min_col=min_column+7,
                        max_col=min_column+12,
                        min_row=min_row+1,
                        max_row=max_row
                        )

        categories = Reference(workbook,
                                min_col=min_column,
                                max_col=min_column,
                                min_row=min_row+3,
                                max_row=max_row
                                )

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categories)


        workbook.add_chart(barchart, 'H12')
        barchart.title = 'Daily Gross Revenue Male'
        barchart.type = 'col'
        barchart.grouping = 'stacked'
        barchart.overlap = 100
        barchart.style = 2


    def total_revenue_perproduct(self, max_column, max_row, min_row, wb):
        alphabet = list(string.ascii_uppercase)
        alphabet_excel = alphabet[:max_column]
        #[A,B,C,D,E,F,G]
        for i in alphabet_excel:
            if i != 'A':
                wb[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
                wb[f'{i}{max_row+1}'].style = 'Currency'

        wb[f'{alphabet_excel[0]}{max_row+1}'] = 'Total Revenue Product'
<<<<<<< HEAD
  
=======
        
    
>>>>>>> d76207fe55e9ee775804835396cdc28e88d91349
    def add_title(self, workbook):
        workbook['A1'] = 'Daily Gross Revenue Report'
        workbook['A2'] = '2019'
        workbook['A1'].font = Font('Calibri', bold=True, size=20)
        workbook['A2'].font = Font('Calibri', bold=True, size=15)

    def save_file(self, wb):
        wb.save(self.output_file)
        print('File saved')
